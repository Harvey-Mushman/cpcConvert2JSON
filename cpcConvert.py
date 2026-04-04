import anthropic
import pdfplumber
import json
import re
from pathlib import Path
from datetime import datetime

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

api_key = Path("anthropicKey.txt").read_text().strip()
client = anthropic.Anthropic(api_key=api_key)

LOG_FILE = Path("./conversion_log.txt")
COUNTY_CONFIGS_FOLDER = Path("./county_configs")
COUNTY_CONFIGS_FOLDER.mkdir(exist_ok=True)

STANDARD_SCHEMA = {
    "certificate_number":          "Unique certificate ID number",
    "issuing_county":              "California county that issued the certificate",
    "issuing_date":                "Date the certificate was issued",
    "expiration_date":             "Date the certificate expires",
    "amended_date":                "Amendment date if any",
    "county_fee":                  "Certificate fee amount",
    "certified_copies_made":       "Number of certified copies issued",
    "producer.name":               "Producer or owner full name",
    "producer.farm_name":          "Farm or business name",
    "producer.dba":                "Doing business as name",
    "producer.address":            "Street address",
    "producer.city":               "City",
    "producer.state":              "State abbreviation",
    "producer.zip_code":           "ZIP code",
    "producer.phone_cell":         "Cell or mobile phone",
    "producer.phone_business":     "Business phone",
    "producer.fax":                "Fax number",
    "producer.email":              "Email address",
    "production_sites":            "Farm/production site list: site_number, address, city, state, zip_code, acreage",
    "storage_locations":           "Storage location list: location_id, address, city, state, zip_code",
    "authorized_counties":         "Counties where produce may be sold",
    "producers_selling_for_me":    "Other CPC holders authorized to sell THIS producer's products: name, certificate_number, date_declared",
    "producers_i_sell_for":        "Producers THIS holder is authorized to sell for: name, certificate_number, date_declared",
    "authorized_representatives":  "Named individuals authorized to operate at this producer's market stall",
    "commodity.site":              "Production site number",
    "commodity.commodity":         "Crop or commodity name",
    "commodity.variety":           "Variety or type",
    "commodity.amount_grown":      "Quantity grown",
    "commodity.est_production":    "Estimated production or yield",
    "commodity.harvest_season":    "Harvest season or months",
    "commodity.season_altering_device": "Season altering device e.g. greenhouse",
    "commodity.months_in_storage": "Storage duration"
}

# ── HELPERS ─────────────────────────────────────────────────────

def normalize(text):
    """Replace internal whitespace/newlines with single spaces and strip."""
    if text is None:
        return ""
    return re.sub(r'\s+', ' ', str(text)).strip()

def clean_json_response(raw):
    raw = raw.strip()
    if raw.startswith("```"):
        parts = raw.split("```")
        raw = parts[1] if len(parts) > 1 else raw
        if raw.startswith("json"):
            raw = raw[4:]
    return raw.strip()

def log(msg, also_print=True):
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(msg + "\n")
    if also_print:
        print(msg)

def detect_form_revision(page_text):
    """Extract state form number and revision from page 1 text, e.g. '51-049 (Rev 10/24)'."""
    m = re.search(r'(5\d-\d{3}[A-Z]?\s*\(Rev\s+[\d/]+\))', page_text, re.IGNORECASE)
    return m.group(1).strip() if m else None

def revision_to_filename_part(revision):
    """Convert '51-049 (Rev 10/24)' to a Windows-safe filename segment '51-049_Rev_10-24'."""
    safe = re.sub(r'[()]+', '', revision)      # remove parentheses
    safe = re.sub(r'/', '-', safe)             # replace / with -
    safe = re.sub(r'\s+', '_', safe).strip('_')  # spaces to underscores
    return safe

def find_county_config(county, form_revision):
    """
    Look up the right config file using a fallback chain:
      1. {COUNTY}_{revision}.json  — exact match for this form version
      2. {COUNTY}.json             — generic / legacy config
    Returns (config_dict, config_path) or (None, None) if not found.
    """
    candidates = []
    if form_revision:
        rev_part = revision_to_filename_part(form_revision)
        candidates.append(COUNTY_CONFIGS_FOLDER / f"{county}_{rev_part}.json")
    candidates.append(COUNTY_CONFIGS_FOLDER / f"{county}.json")

    for path in candidates:
        if path.exists():
            return json.loads(path.read_text()), path
    return None, None

def detect_county(page_text, tables):
    """Find issuing county from page 1 text or table data."""
    patterns = [
        r'Issuing County[:\s]+([A-Za-z][A-Za-z\s]+?)(?:\n|$)',
        r'ISSUING COUNTY[:\s]+([A-Z][A-Z\s]+?)(?:\n|$)',
    ]
    for pattern in patterns:
        m = re.search(pattern, page_text, re.IGNORECASE)
        if m:
            return m.group(1).strip().upper()
    # Try tables
    for table in tables:
        for row in table:
            for cell in row:
                if cell:
                    m = re.search(r'Issuing County[:\s]+([A-Za-z][A-Za-z\s]+)', normalize(cell), re.IGNORECASE)
                    if m:
                        return m.group(1).strip().upper()
    return None

def get_table_name(table):
    """First cell of first row, normalized."""
    if table and table[0] and table[0][0]:
        return normalize(table[0][0])
    return None

def is_page_repeat(table_name, config):
    """True if this table is a per-page repeat summary to skip."""
    for prefix in config.get("page_repeat_prefixes", ["Certified Producer:"]):
        if table_name.startswith(prefix):
            return True
    return False

def find_header_row(table):
    """Find the index of the row that contains column headers (first row with 2+ values)."""
    for idx, row in enumerate(table):
        if idx == 0:
            continue
        filled = [c for c in row if c and normalize(c)]
        if len(filled) >= 2:
            return idx
    return 1

# ── TABLE PROCESSORS ────────────────────────────────────────────

def process_key_value(table, table_cfg):
    """Two-column table: left=label, right=value. Row 0 is title."""
    field_map = table_cfg.get("field_map", {})
    result = {}
    for row in table[1:]:
        if len(row) >= 2:
            label = normalize(row[0])
            value = normalize(row[1])
            if label:
                standard_key = field_map.get(label, label)
                result[standard_key] = value
    return result

def process_data_table(table, table_cfg):
    """Title row, header row, data rows.
    Use header_row=0 in config when the table has no separate title — row 0 IS the header."""
    col_map = table_cfg.get("column_map", {})
    forced_header = table_cfg.get("header_row", None)
    if forced_header is not None:
        header_idx = forced_header
    else:
        header_idx = find_header_row(table)
    if header_idx >= len(table):
        return []
    headers = [col_map.get(normalize(h), normalize(h)) for h in table[header_idx]]
    rows = []
    for row in table[header_idx + 1:]:
        if not any(c and normalize(c) for c in row):
            continue
        entry = {h: normalize(v) for h, v in zip(headers, row) if h}
        rows.append(entry)
    return rows

def process_list(table, table_cfg):
    """Single-column or comma-separated list.
    Use skip_title_row=false when row 0 contains data (no separate title row)."""
    start = 0 if table_cfg.get("skip_title_row") is False else 1
    items = []
    for row in table[start:]:
        for cell in row:
            if cell and normalize(cell):
                for item in normalize(cell).split(","):
                    item = item.strip()
                    if item:
                        items.append(item)
    return items

def process_paired_sites(table, table_cfg):
    """Production site table where each row holds 2 sites in alternating (description, acreage) columns.
    The table name row (row 0) is the first site — no separate title row."""
    desc_pattern = table_cfg.get("desc_pattern", r"^(\d+)\.?\s*(.+)$")
    sites = []
    for row in table:             # start at row 0 — it IS data
        for pair_start in range(0, len(row) - 1, 2):
            desc_cell  = normalize(row[pair_start])     if pair_start < len(row) else ""
            acres_cell = normalize(row[pair_start + 1]) if pair_start + 1 < len(row) else ""
            if not desc_cell:
                continue
            m = re.match(desc_pattern, desc_cell)
            if m:
                sites.append({
                    "site_number":  m.group(1).strip(),
                    "description":  m.group(2).strip(),
                    "acreage":      acres_cell
                })
    return sites

def process_location_cells(table, table_cfg):
    """Storage location table where each cell contains 'Storage Location (X): address'.
    The table name row IS the first data row."""
    cell_pattern = table_cfg.get("cell_pattern", r"Storage Location\s*\(([^)]+)\)\s*[:\s]+(.+)")
    locations = []
    for row in table:             # start at row 0 — it IS data
        for cell in row:
            if not cell:
                continue
            m = re.search(cell_pattern, normalize(cell), re.IGNORECASE)
            if m:
                locations.append({
                    "location_id": m.group(1).strip(),
                    "address":     m.group(2).strip()
                })
    return locations

def process_split_data_table(table, table_cfg):
    """Table where left columns map to one field and right columns map to another.
    Example: Kern County combines producers_i_sell_for and authorized_representatives side by side."""
    result = {}
    for side in ("left", "right"):
        side_cfg  = table_cfg.get(side, {})
        cols      = side_cfg.get("columns", [])
        col_map   = side_cfg.get("column_map", {})
        maps_to   = side_cfg.get("maps_to", side)
        skip_vals = {v.upper() for v in side_cfg.get("skip_values", [])}
        if not cols:
            continue
        # Row 0 contains the column headers for this side
        headers = [col_map.get(normalize(table[0][c]), normalize(table[0][c]))
                   if c < len(table[0]) else ""
                   for c in cols]
        rows = []
        for row in table[1:]:
            values = [normalize(row[c]) if c < len(row) else "" for c in cols]
            if all(v.upper() in skip_vals or not v for v in values):
                continue
            entry = {h: v for h, v in zip(headers, values) if h}
            rows.append(entry)
        result[maps_to] = rows
    return result

def process_certificate_fields(table, table_cfg):
    """Borderless label:value table for certificate header."""
    field_map = table_cfg.get("field_map", {})
    result = {}
    for row in table[1:]:
        if len(row) >= 2:
            label = normalize(row[0])
            value = normalize(row[1])
            standard_key = field_map.get(label, None)
            if standard_key and value:
                result[standard_key] = value
    return result

def process_049m_producer_from_page(page, table_cfg):
    """Parse Tulare-style producer table using char-level extraction from the page.
    The producer table has values and labels at specific x-positions."""
    producer = {}
    chars = page.chars
    from collections import defaultdict

    # Group chars into lines by y position
    buckets = defaultdict(list)
    for c in chars:
        buckets[round(c['top'] / 12) * 12].append(c)

    def get_text_at(y_key, x_min, x_max):
        line_chars = [c for c in buckets.get(y_key, []) if c['x0'] >= x_min and c['x0'] < x_max]
        return ''.join(c['text'] for c in sorted(line_chars, key=lambda c: c['x0'])).strip()

    def get_groups(y_key, gap=8):
        line_chars = sorted(buckets.get(y_key, []), key=lambda c: c['x0'])
        groups = []
        current = ""
        current_x0 = None
        prev_x1 = None
        for ch in line_chars:
            g = ch['x0'] - prev_x1 if prev_x1 is not None else 0
            if g > gap and current:
                groups.append((current_x0, current.strip()))
                current = ch['text']
                current_x0 = ch['x0']
            else:
                if current_x0 is None:
                    current_x0 = ch['x0']
                current += ch['text']
            prev_x1 = ch['x1']
        if current:
            groups.append((current_x0, current.strip()))
        return groups

    # Find key label lines by searching for known text (first occurrence only)
    label_ys = {}
    for y_key in sorted(buckets.keys()):
        text = ''.join(c['text'] for c in sorted(buckets[y_key], key=lambda c: c['x0']))
        if "Name of Producer" in text and "name_label" not in label_ys:
            label_ys["name_label"] = y_key
        if "Farm or Ranch Name" in text and "farm_label" not in label_ys:
            label_ys["farm_label"] = y_key
        if "Mailing Address" in text and "address_label" not in label_ys:
            label_ys["address_label"] = y_key
        if "Phone Number" in text and ("Alt" in text or "2nd" in text) and "phone_label" not in label_ys:
            label_ys["phone_label"] = y_key

    sorted_ys = sorted(buckets.keys())

    # Producer name: line before "Name of Producer" label
    if "name_label" in label_ys:
        prev_y = max(y for y in sorted_ys if y < label_ys["name_label"])
        groups = get_groups(prev_y)
        producer["name"] = groups[0][1] if groups else ""

    # Farm name + DBA: look for values between name_label and farm_label lines
    if "name_label" in label_ys and "farm_label" in label_ys:
        # First check right side of name_label line (Tulare style: name and farm on same line)
        groups = get_groups(label_ys["name_label"])
        farm_parts = [g[1] for g in groups if g[0] > 300]
        # Then check lines between name_label and farm_label (SLO style: farm on its own line)
        for y_key in sorted_ys:
            if label_ys["name_label"] < y_key < label_ys["farm_label"]:
                cont_groups = get_groups(y_key)
                for g in cont_groups:
                    farm_parts.append(g[1])
        farm_text = " ".join(farm_parts).strip().strip("/").strip()
        if "/" in farm_text:
            parts = farm_text.split("/", 1)
            producer["farm_name"] = parts[0].strip()
            producer["dba"] = parts[1].strip()
        else:
            producer["farm_name"] = farm_text

    # Address/City/State/Zip: line before "Mailing Address" label
    if "address_label" in label_ys:
        prev_y = max(y for y in sorted_ys if y < label_ys["address_label"])
        groups = get_groups(prev_y)
        if len(groups) >= 4:
            producer["address"] = groups[0][1]
            producer["city"] = groups[1][1]
            producer["state"] = groups[2][1]
            producer["zip_code"] = groups[3][1]
        elif len(groups) >= 1:
            producer["address"] = " ".join(g[1] for g in groups)

    # Phone/2nd Phone/Email: line before "Phone Number" label
    if "phone_label" in label_ys:
        prev_y = max(y for y in sorted_ys if y < label_ys["phone_label"])
        groups = get_groups(prev_y)
        if len(groups) >= 3:
            producer["phone_number"] = groups[0][1]
            producer["2nd_phone_number"] = groups[1][1]
            producer["email"] = groups[2][1]
        elif len(groups) == 2:
            producer["phone_number"] = groups[0][1]
            if "@" in groups[1][1]:
                producer["email"] = groups[1][1]
            else:
                producer["2nd_phone_number"] = groups[1][1]
        elif len(groups) == 1:
            producer["phone_number"] = groups[0][1]

    # Ensure all standard producer fields are present
    for field in ("name", "farm_name", "dba", "address", "city", "state",
                  "zip_code", "phone_number", "2nd_phone_number", "email"):
        producer.setdefault(field, "")

    return producer

def process_049m_production_sites(table, table_cfg):
    """Parse 51-049M production site table.
    Each row is one cell with two formats:
      Tulare: 'ADDRESS, CITY ACREAGE\\nAddress of Production Site N [DATE] Site Acreage'
      SLO:    'N ADDRESS CITY STATE ZIP ACREAGE\\nProduction Site N Address Site Acreage'"""
    sites = []
    for row in table[1:]:
        if not row or not row[0]:
            continue
        cell = normalize(row[0])
        # Try "Address of Production Site N" (Tulare)
        m = re.search(r'Address of Production Site\s+(\d+)', cell)
        # Try "Production Site N Address" (SLO)
        if not m:
            m = re.search(r'Production Site\s+(\d+)\s+Address', cell)
        if not m:
            continue
        site_num = m.group(1)
        # Text before the label line contains the data
        label_start = cell.find("Address of Production Site")
        if label_start == -1:
            label_start = cell.find("Production Site")
        before = cell[:label_start].strip() if label_start > 0 else ""
        # If data starts with the site number, strip it (SLO format: "1 2299 Bonita...")
        before = re.sub(r'^\d+\s+', '', before, count=1)
        # Last token(s) may be acreage
        acreage_m = re.search(r'([\d.]+)\s*$', before)
        acreage = acreage_m.group(1) if acreage_m else ""
        address = before[:acreage_m.start()].strip().rstrip(",") if acreage_m else before
        sites.append({
            "site_number": site_num,
            "address": address,
            "acreage": acreage
        })
    return sites

def process_049m_storage(table, table_cfg):
    """Parse Tulare-style storage location table.
    Each row is one cell: 'ADDRESS\\nStorage Location (A)'"""
    locations = []
    for row in table[1:]:
        if not row or not row[0]:
            continue
        cell = row[0]
        m = re.search(r'Storage Location\s*\(([A-Z])\)', cell)
        if not m:
            continue
        loc_id = m.group(1)
        address = cell[:m.start()].replace("\n", " ").strip()
        locations.append({"location_id": loc_id, "address": address})
    return locations

def process_049m_commodities(table, table_cfg):
    """Parse Tulare-style commodity table.
    Row 0 = title, Row 1 = header, Rows 2+ = space-separated data in single cells.
    Format: SITE COMMODITY [VARIETY] AMOUNT UNIT EST_PROD UNIT SEASON DEVICE MONTHS
    Variety may be missing or multi-word. Parse from right (fixed fields) then left."""
    commodities = []
    for row in table[2:]:  # skip title + header
        if not row or not row[0]:
            continue
        cell = normalize(row[0])
        if not cell or not cell[0].isdigit():
            continue

        # Parse from right: MONTHS DEVICE SEASON EST_UNIT EST_NUM AMT_UNIT AMT_NUM
        # Then remaining text = SITE COMMODITY [VARIETY]
        m = re.match(
            r'^(\d{1,2})\s+'                        # site
            r'(.+?)\s+'                              # commodity + variety (greedy middle)
            r'([\d.]+)\s+'                           # amount number
            r'([A-Z]+)\s+'                           # amount unit
            r'([\d.]+)\s+'                           # est_production number
            r'([A-Z]+)\s+'                           # est_production unit
            r'([A-Z]{3,}-[A-Z]{3,}|ALL\s+YEAR)\s+'  # harvest season
            r'(\S+)\s+'                              # season_altering_device
            r'(\d+)'                                 # months_in_storage
            r'(.*)$',                                # trailing text (variety overflow)
            cell
        )
        if m:
            site = m.group(1)
            name_block = m.group(2).strip()
            trailing = m.group(10).strip()

            # Split commodity from variety: first word is commodity, rest is variety
            # But commodity can be multi-word (e.g. "FIG BLACK")
            # Heuristic: commodity is the first word, variety is the rest
            # Unless there's no variety (single word like "BLACKBERRY")
            words = name_block.split()
            commodity = words[0] if words else ""
            variety = " ".join(words[1:]) if len(words) > 1 else ""

            # If there's trailing text after months_in_storage, it's variety overflow
            if trailing:
                variety = (variety + " " + trailing).strip() if variety else trailing

            # Clean variety of wrapping artifacts
            variety = re.sub(r'[()]', '', variety).strip().strip(',').strip()

            device = m.group(8)
            if device.upper() == "NONE":
                device = ""

            commodities.append({
                "site": site,
                "commodity": commodity,
                "variety": variety,
                "amount_grown": f"{m.group(3)} {m.group(4)}",
                "est_production": f"{m.group(5)} {m.group(6)}",
                "harvest_season": m.group(7),
                "season_altering_device": device,
                "months_in_storage": m.group(9)
            })
    return commodities

def process_049m_second_certs(table, table_cfg):
    """Parse Tulare-style second certificates table.
    Rows contain sell-for and sell-for-me side by side.
    All rows are included even if values are N/A."""
    result = {"producers_i_sell_for": [], "producers_selling_for_me": []}
    # Find the data rows (after header rows)
    data_start = 0
    for i, row in enumerate(table):
        if row and row[0] and "Name of Producer" in str(row[0]):
            data_start = i + 1
            break
    for row in table[data_start:]:
        if not row or len(row) < 8:
            continue
        # Left side: sell for (cols 0, 2, 3)
        name = normalize(row[0]) if row[0] else ""
        cert = normalize(row[2]) if len(row) > 2 and row[2] else ""
        date = normalize(row[3]) if len(row) > 3 and row[3] else ""
        if name:
            result["producers_i_sell_for"].append({
                "name": name, "certificate_number": cert, "date_declared": date
            })
        # Right side: sell for me (cols 5, 6, 7)
        name2 = normalize(row[5]) if len(row) > 5 and row[5] else ""
        cert2 = normalize(row[6]) if len(row) > 6 and row[6] else ""
        date2 = normalize(row[7]) if len(row) > 7 and row[7] else ""
        if name2:
            result["producers_selling_for_me"].append({
                "name": name2, "certificate_number": cert2, "date_declared": date2
            })
    return result

def process_049m_auth_counties(table, table_cfg):
    """Parse Tulare-style authorized counties table (multi-column with parenthetical numbers)."""
    counties = []
    for row in table[1:]:  # skip title
        for cell in row:
            if not cell:
                continue
            name = normalize(cell).strip()
            if name and name != "SECOND CERTIFICATES":
                counties.append(name)
    return counties

def process_049m_auth_reps(table, table_cfg):
    """Parse Tulare-style authorized representatives table.
    Returns objects with last_name, first_name, phone_number."""
    reps = []
    for row in table[2:]:  # skip title + header
        if not row:
            continue
        last = normalize(row[0]) if row[0] else ""
        first = normalize(row[1]) if len(row) > 1 and row[1] else ""
        phone = normalize(row[2]) if len(row) > 2 and row[2] else ""
        if last or first:
            reps.append({
                "last_name": last,
                "first_name": first,
                "phone_number": phone
            })
    return reps

def process_049m_mega_table(table, table_cfg):
    """Process a merged table that contains COMMODITIES + AUTHORIZED COUNTIES +
    SECOND CERTIFICATES in one pdfplumber table (51-049M form).
    Splits the table at section header rows and processes each section."""
    result = {}

    # Find section boundaries by looking for header rows
    sections = []  # list of (section_name, start_row, end_row)
    for i, row in enumerate(table):
        first_cell = normalize(row[0]) if row and row[0] else ""
        if first_cell.startswith("COMMODITIES"):
            sections.append(("commodities", i, None))
        elif first_cell == "AUTHORIZED COUNTIES":
            sections.append(("authorized_counties", i, None))
        elif first_cell == "SECOND CERTIFICATES":
            sections.append(("second_certificates", i, None))
        elif first_cell == "AUTHORIZED REPRESENTATIVES":
            sections.append(("authorized_representatives", i, None))

    # Set end rows
    for idx in range(len(sections)):
        if idx + 1 < len(sections):
            sections[idx] = (sections[idx][0], sections[idx][1], sections[idx + 1][1])
        else:
            sections[idx] = (sections[idx][0], sections[idx][1], len(table))

    for section_name, start, end in sections:
        sub_table = table[start:end]

        if section_name == "commodities":
            # Row 0 = title ("COMMODITIES - N Total Lines"), Row 1 = headers, Row 2+ = data
            col_map = table_cfg.get("commodity_column_map", {})
            if len(sub_table) > 1:
                # Build header map from row 1
                headers = sub_table[1]
                mapped_headers = []
                for h in headers:
                    h_norm = normalize(h) if h else ""
                    mapped_headers.append(col_map.get(h_norm, h_norm))

                commodities = []
                for row in sub_table[2:]:
                    if not any(c and normalize(c) for c in row):
                        continue
                    entry = {}
                    for h, v in zip(mapped_headers, row):
                        if h:
                            entry[h] = normalize(v) if v else ""
                    # Split "commodity, variety" if combined
                    if "commodity" in entry and "variety" not in entry:
                        comm_val = entry["commodity"]
                        if "," in comm_val:
                            parts = comm_val.split(",", 1)
                            entry["commodity"] = parts[0].strip()
                            entry["variety"] = parts[1].strip()
                        else:
                            entry["variety"] = ""
                    for std in ("site", "commodity", "variety", "amount_grown",
                                "est_production", "harvest_season",
                                "season_altering_device", "months_in_storage"):
                        entry.setdefault(std, "")
                    commodities.append(entry)
                if commodities:
                    result.setdefault("commodities", []).extend(commodities)

        elif section_name == "authorized_counties":
            # Row 0 = "AUTHORIZED COUNTIES", remaining rows have county names in cells
            counties = []
            for row in sub_table[1:]:
                for cell in row:
                    if cell and normalize(cell):
                        name = re.sub(r'\s*\(\d+\)\s*$', '', normalize(cell)).strip()
                        if name:
                            counties.append(name)
            result["authorized_counties"] = counties

        elif section_name == "second_certificates":
            # Same structure as Tulare second certs
            data = process_049m_second_certs(sub_table, table_cfg)
            for key, rows in data.items():
                result.setdefault(key, []).extend(rows)

        elif section_name == "authorized_representatives":
            reps = process_049m_auth_reps(sub_table, table_cfg)
            result.setdefault("authorized_representatives", []).extend(reps)

    return result

def extract_cert_fields_from_text(page_text, config):
    """Extract certificate fields from page text using regex.
    Stops at end of line; skips empty values."""
    result = {}
    for label, standard_key in config.get("certificate_fields_from_text", {}).items():
        # Use [: \t]+ so we never cross a newline when the field is empty
        pattern = re.escape(label) + r'[: \t]+([^\n]*)'
        m = re.search(pattern, page_text, re.IGNORECASE)
        if m:
            value = m.group(1).strip()
            if value:
                result[standard_key] = value
    return result

def _build_label_stop_pattern(config):
    """Build a regex lookahead that stops before any known field label.
    Labels that already end with a non-word character (e.g. 'PHONE: (cell)')
    are matched as-is; plain word labels get a trailing \\s*: added.
    Additional stop labels (e.g. 'ZIP CODE') can be listed in config under
    'additional_stop_labels' to catch adjacent-column bleed without extracting them."""
    labels = (list(config.get("certificate_fields_from_text", {}).keys()) +
              list(config.get("producer_fields_from_text", {}).keys()) +
              config.get("additional_stop_labels", []))
    if not labels:
        return None
    labels.sort(key=len, reverse=True)
    parts = []
    for l in labels:
        escaped = re.escape(l)
        # If the label ends with a word character, it needs a colon to confirm it's a label
        if re.search(r'\w$', l):
            parts.append(escaped + r'\s*:')
        else:
            parts.append(escaped)
    return r'(?=' + '|'.join(parts) + r'|\n|$)'

def extract_producer_fields_from_text(page_text, config):
    """Extract producer fields from page text using regex.
    Stops at the next known field label on the same line to avoid grabbing
    two-column values. If a label has no value on the same line (e.g. DBA:
    followed by newline), checks the next line for the value."""
    field_map = config.get("producer_fields_from_text", {})
    if not field_map:
        return {}
    stop = _build_label_stop_pattern(config) or r'(?=\n|$)'
    producer = {}
    for label, field_name in field_map.items():
        pattern = re.escape(label) + r'[: \t]+([^\n]*?)' + stop
        m = re.search(pattern, page_text, re.IGNORECASE)
        if m:
            value = m.group(1).strip()
            if value:
                producer[field_name] = value
            else:
                # Value is empty on same line — check next line
                label_pos = m.start()
                newline_pos = page_text.find('\n', label_pos)
                if newline_pos >= 0:
                    next_newline = page_text.find('\n', newline_pos + 1)
                    if next_newline == -1:
                        next_newline = len(page_text)
                    next_line = page_text[newline_pos + 1:next_newline].strip()
                    # Only use next line if it's not another label
                    if next_line and not any(next_line.startswith(l) for l in field_map.keys()):
                        producer[field_name] = next_line
    return producer

def extract_text_sections(page_text, config):
    """Parse sections from page text using start/end anchors."""
    result = {}
    for section_name, section_cfg in config.get("text_sections", {}).items():
        start_marker = section_cfg.get("start", "")
        end_marker   = section_cfg.get("end", "")
        maps_to      = section_cfg.get("maps_to", section_name)
        parser       = section_cfg.get("parser", "lines")
        skip_text    = section_cfg.get("skip_if_text", "")

        # Find the section in text
        start_idx = page_text.find(start_marker)
        if start_idx == -1:
            continue
        start_idx += len(start_marker)

        if end_marker:
            end_idx = page_text.find(end_marker, start_idx)
            section_text = page_text[start_idx:end_idx].strip() if end_idx != -1 else page_text[start_idx:].strip()
        else:
            section_text = page_text[start_idx:].strip()

        if not section_text:
            continue
        if skip_text and section_text.upper().startswith(skip_text.upper()):
            continue

        if parser == "comma_list":
            items = [i.strip() for i in re.split(r'[,\n]+', section_text) if i.strip()]
            result[maps_to] = items
        elif parser == "lines":
            items = [ln.strip() for ln in section_text.splitlines() if ln.strip()]
            result[maps_to] = items
        elif parser == "tabular_lines":
            columns = section_cfg.get("columns", [])
            rows = []
            for line in section_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = re.split(r'\s{2,}', line)
                entry = {}
                for i, col in enumerate(columns):
                    entry[col] = parts[i].strip() if i < len(parts) else ""
                rows.append(entry)
            result[maps_to] = rows
        elif parser == "regex_lines":
            # Parse each line with a regex, mapping capture groups to field names
            line_pattern = section_cfg.get("line_pattern", "")
            columns = section_cfg.get("columns", [])
            skip_header = section_cfg.get("skip_header_rows", 0)
            rows = []
            lines = section_text.splitlines()
            for line in lines[skip_header:]:
                line = line.strip()
                if not line:
                    continue
                m = re.match(line_pattern, line)
                if m:
                    entry = {}
                    for i, col in enumerate(columns):
                        val = m.group(i + 1).strip() if i + 1 <= len(m.groups()) else ""
                        if col != "_skip":
                            entry[col] = val
                    rows.append(entry)
            if rows:
                result.setdefault(maps_to, []).extend(rows)
        elif parser == "numbered_sites":
            # Parse "1. addr 2. addr 3. 4." inline numbered format
            # Use regex to find "N. text" where text is a real address (contains letters)
            flat = re.sub(r'\s+', ' ', section_text.replace("\n", " ")).strip()
            matches = re.findall(r'(\d{1,2})\.\s+((?:(?!\s\d{1,2}\.).)+)', flat)
            sites = []
            for num, addr in matches:
                addr = addr.strip().rstrip(",. ").strip()
                # Skip if addr is just digits or empty (empty site slot)
                if addr and re.search(r'[a-zA-Z]', addr):
                    sites.append({"site_number": num, "description": addr})
            if sites:
                result[maps_to] = sites
        elif parser == "lettered_sites":
            # Parse "A. addr B. addr C. addr D." inline lettered format
            # All listed letters are included, even if address is empty
            flat = re.sub(r'\s+', ' ', section_text.replace("\n", " ")).strip()
            # Split on letter-dot pattern (A-D only — CPC forms have max 4 slots)
            parts = re.split(r'\b([A-D])\.\s*', flat)
            locations = []
            # parts: ['', 'A', 'addr...', 'B', 'addr or empty', ...]
            for i in range(1, len(parts) - 1, 2):
                letter = parts[i]
                addr = parts[i + 1].strip().rstrip(",. ").strip() if i + 1 < len(parts) else ""
                locations.append({"location_id": letter, "address": addr})
            if locations:
                result[maps_to] = locations
    return result

def process_production_sites_cells(table, table_pattern_cfg):
    """Parse San Diego-style production site cells using regex."""
    pattern  = table_pattern_cfg.get("cell_pattern", "")
    fields   = table_pattern_cfg.get("fields", [])
    sites = []
    for row in table:
        for cell in row:
            if not cell:
                continue
            cell_text = normalize(cell)
            m = re.search(pattern, cell_text, re.IGNORECASE)
            if m:
                entry = {}
                for i, field_name in enumerate(fields, start=1):
                    entry[field_name] = m.group(i).strip() if i <= len(m.groups()) else ""
                sites.append(entry)
    return sites

def process_positional_table(table, table_pattern_cfg):
    """Parse a table by column position rather than header names."""
    col_map   = table_pattern_cfg.get("columns", {})
    skip_rows = table_pattern_cfg.get("skip_rows", 1)
    rows = []
    for row in table[skip_rows:]:
        if not any(c and normalize(c) for c in row):
            continue
        entry = {}
        for pos_str, field_name in col_map.items():
            pos = int(pos_str)
            entry[field_name] = normalize(row[pos]) if pos < len(row) else ""
        rows.append(entry)
    return rows

# ── KERN DUAL-COLUMN HEADER EXTRACTION ───────────────────────────

def _deinterleave_by_text(row_chars, ref_chars, x_tol=1.5):
    """Remove chars from row_chars that match a ref char by same text + close x.
    Returns the remaining chars (the left-column stream)."""
    ref_used = set()
    ref_sorted = sorted(enumerate(ref_chars), key=lambda t: t[1]['x0'])
    unique = []
    for c in sorted(row_chars, key=lambda c: c['x0']):
        matched = False
        for i, rc in ref_sorted:
            if i in ref_used:
                continue
            if abs(c['x0'] - rc['x0']) <= x_tol and c['text'] == rc['text']:
                ref_used.add(i)
                matched = True
                break
        if not matched:
            unique.append(c)
    return unique

def _deinterleave_by_position(row_chars, ref_chars, x_tol=2.0):
    """Remove chars from row_chars whose x position is closest to a ref char (no text check).
    Returns the remaining chars (the left-column stream)."""
    remove_idx = set()
    row_sorted = list(enumerate(sorted(row_chars, key=lambda c: c['x0'])))
    for rc in sorted(ref_chars, key=lambda c: c['x0']):
        best_i, best_delta = None, float('inf')
        for i, c in row_sorted:
            if i in remove_idx:
                continue
            delta = abs(c['x0'] - rc['x0'])
            if delta < best_delta and delta <= x_tol:
                best_delta = delta
                best_i = i
        if best_i is not None:
            remove_idx.add(best_i)
    return [c for i, c in row_sorted if i not in remove_idx]

def extract_kern_official_header(page):
    """Extract certificate fields from Kern County's FOR OFFICIAL USE ONLY table.

    The PDF renders a 3-row × 2-column grid (left=cert fields, right=expiration fields)
    at the same y coordinates, causing pdfplumber to interleave characters from both
    columns into each row.  The right-column values also appear a second time in clean
    rows below (rows 4-6), which we use as de-interleaving references.

    Row pairing:
      garbled row 1 (CERTIFICATE NO)   ←→ clean row 4 (EXPIRATION DATE)
      garbled row 2 (COUNTY FEE)       ←→ clean row 5 (AMENDED DATE – empty)
      garbled row 3 (ISSUING DATE)     ←→ clean row 6 (CERTIFIED COPIES MADE)
    """
    result = {}

    # Find the FOR OFFICIAL USE ONLY table object (need its bbox for char-level crop)
    foo_table = None
    for t in page.find_tables():
        cells = t.extract()
        if cells and cells[0] and cells[0][0]:
            first = (cells[0][0] or '').strip()
            if first.startswith('FOR OFFICIAL USE ONLY'):
                foo_table = t
                break
    if not foo_table:
        return result

    region = page.crop(foo_table.bbox)
    chars = region.chars

    # Group chars into rows (bucket by y, skip the header band at y≈54)
    from collections import defaultdict, Counter
    buckets = defaultdict(list)
    for c in chars:
        buckets[round(c['top'] / 20) * 20].append(c['top'])
    row_ys = sorted(
        [sum(v) / len(v) for v in buckets.values() if sum(v) / len(v) > 60]
    )

    if len(row_ys) < 6:
        log(f"    WARNING: extract_kern_official_header found only {len(row_ys)} rows, expected 6")
        return result

    # Detect value font from the expiration-date row (row 4), which has a clean label+value.
    # The label ends at roughly table_x0+85 and the value starts after that.
    # Using the right half (x > table midpoint) avoids the label entirely.
    mid_x_page = (foo_table.bbox[0] + foo_table.bbox[2]) / 2
    exp_y_approx = row_ys[3]
    exp_right = [c for c in chars
                 if abs(c['top'] - exp_y_approx) < 5
                 and c['x0'] > mid_x_page
                 and c['text'].strip()]
    right_fonts = Counter(c['fontname'] for c in exp_right)
    val_font = right_fonts.most_common(1)[0][0] if right_fonts else 'CIDFont+F6'

    y_cert, y_fee, y_issue, y_exp, y_amend, y_copies = row_ys[:6]

    def get_row(y_target, y_tol=5):
        return sorted(
            [c for c in chars if abs(c['top'] - y_target) < y_tol and c['fontname'] == val_font],
            key=lambda c: c['x0']
        )

    def txt(cl):
        return ''.join(c['text'] for c in cl).strip()

    # Row 2 (COUNTY FEE): amended_date is empty so no interleaving — read directly
    result['county_fee'] = txt(get_row(y_fee))

    # Row 4 (EXPIRATION DATE): clean, read directly
    result['expiration_date'] = txt(get_row(y_exp))

    # Row 5 (AMENDED DATE): clean, read directly — always include even if empty
    result['amended_date'] = txt(get_row(y_amend))

    # Row 6 (CERTIFIED COPIES MADE): clean, read directly
    result['certified_copies_made'] = txt(get_row(y_copies))

    # Row 1 (CERTIFICATE NO): subtract expiration-date chars by same-text + close-x
    cert_chars = _deinterleave_by_text(get_row(y_cert), get_row(y_exp))
    result['certificate_number'] = txt(cert_chars)

    # Row 3 (ISSUING DATE): subtract certified-copies chars by position only
    issue_chars = _deinterleave_by_position(get_row(y_issue), get_row(y_copies))
    result['issuing_date'] = txt(issue_chars)

    # Keep empty strings (e.g. amended_date when blank) — only drop None
    return {k: v for k, v in result.items() if v is not None}


def find_continuation_config(table, tables_config):
    """Detect if a table is a page-break continuation of a known data_table config.

    Checks two ways:
      1. First cell matches a known table name exactly (title row repeated on new page).
         In this case the table can be processed normally (header row auto-detected).
      2. First row contains column headers that belong to a known data_table.
         In this case header_row=0 must be forced.

    Returns (config_name, table_cfg, force_header_row_0) or (None, None, False).
    """
    if not table or not table[0]:
        return None, None, False

    first_cell = normalize(table[0][0]) if table[0][0] else ""

    # Pass 1: first cell is a known table name → process normally
    if first_cell in tables_config:
        tcfg = tables_config[first_cell]
        if tcfg.get("type") == "data_table":
            return first_cell, tcfg, False

    # Pass 2: first row looks like column headers for a known table
    first_row_cells = [normalize(c) for c in table[0] if c and normalize(c)]
    if first_row_cells:
        for cfg_name, tcfg in tables_config.items():
            if tcfg.get("type") != "data_table":
                continue
            col_map = tcfg.get("column_map", {})
            if not col_map:
                continue
            matches = sum(1 for cell in first_row_cells if cell in col_map)
            if matches >= 2 and matches >= len(first_row_cells) // 2:
                return cfg_name, tcfg, True

    return None, None, False


def _apply_continuation(table, cont_cfg, force_h0, maps_to, result):
    """Process a continuation table and extend result[maps_to]."""
    cfg_use = {**cont_cfg, "header_row": 0} if force_h0 else cont_cfg
    data = process_data_table(table, cfg_use)
    if data:
        result.setdefault(maps_to, []).extend(data)
    return data


def match_table_pattern(table, pattern_cfg):
    """Return True if this table matches the detection rule."""
    detection = pattern_cfg.get("detection", "")
    if not table or not table[0]:
        return False
    first_cell = normalize(table[0][0]) if table[0][0] else ""
    if detection == "first_cell_regex":
        return bool(re.search(pattern_cfg.get("pattern", ""), first_cell))
    elif detection == "first_cell_contains":
        return pattern_cfg.get("contains", "") in first_cell
    return False

# ── PAGE EXTRACTION ──────────────────────────────────────────────

def extract_commodities_by_char_position(page, config):
    """Extract commodity data from free-text lines using character x-positions.
    Used for Santa Barbara and similar counties where commodity data is in
    text lines with positional columns that pdfplumber collapses to single spaces."""
    col_boundaries = config.get("commodity_column_boundaries", [])
    if not col_boundaries:
        return []

    chars = page.chars
    # Identify commodity lines: find ':' chars at x < 45, y > min_y
    min_y = config.get("commodity_min_y", 300)
    colon_chars = sorted(
        [c for c in chars if c['text'] == ':' and c['x0'] < 45 and c['top'] > min_y],
        key=lambda c: c['top']
    )

    commodities = []
    for cc in colon_chars:
        y = cc['top']
        line_chars = sorted(
            [ch for ch in chars if abs(ch['top'] - y) < 3],
            key=lambda ch: ch['x0']
        )

        # Check this is a commodity line (starts with digits + colon)
        prefix = ''.join(ch['text'] for ch in line_chars if ch['x0'] < cc['x1'] + 3)
        if not re.match(r'\d+:', prefix.strip()):
            continue

        # Extract text at each column boundary
        entry = {}
        for col_cfg in col_boundaries:
            field = col_cfg["field"]
            x_min = col_cfg["x_min"]
            x_max = col_cfg["x_max"]
            col_chars = [ch for ch in line_chars if ch['x0'] >= x_min and ch['x0'] < x_max]
            value = ''.join(ch['text'] for ch in col_chars).strip()
            if field != "_skip":
                entry[field] = value

        # Ensure all standard commodity fields present
        for std_field in ("site", "commodity", "variety", "amount_grown",
                          "est_production", "harvest_season",
                          "season_altering_device", "months_in_storage"):
            entry.setdefault(std_field, "")
        commodities.append(entry)

    return commodities

def extract_page_table_based(page, county_config):
    """Standard table-based extraction (most counties)."""
    result = {}
    page_text = page.extract_text() or ""
    tables = page.extract_tables()

    # Log all table names found on this page for diagnostics
    table_names_found = [get_table_name(t) for t in tables if t]
    log(f"    Tables: {table_names_found}", also_print=False)

    # Extract certificate header fields from page text (handles issuing_county for all counties,
    # plus expiration_date/amended_date/etc. for counties that put them in text).
    cert_fields = extract_cert_fields_from_text(page_text, county_config)
    if cert_fields:
        result.update(cert_fields)

    # Kern County dual-column header: the FOR OFFICIAL USE ONLY section interleaves two
    # text streams.  Call the de-interleaving extractor which overrides the fields that
    # would otherwise be garbled (certificate_number, county_fee, issuing_date, etc.).
    if county_config.get("certificate_header_type") == "kern_dual_column":
        header_fields = extract_kern_official_header(page)
        if header_fields:
            result.update(header_fields)

    # Some counties (e.g. Kern) have producer info in free text even though
    # the rest of the document uses tables.
    producer = extract_producer_fields_from_text(page_text, county_config)
    if producer:
        result["producer"] = producer

    tables_config = county_config.get("tables", {})

    for table in tables:
        if not table:
            continue
        table_name = get_table_name(table)
        if not table_name:
            continue
        if is_page_repeat(table_name, county_config):
            log(f"    SKIP page-repeat: '{table_name}' ({len(table)} rows)", also_print=False)
            # If pdfplumber merged the repeat header with continuation rows from the
            # previous page's incomplete table, the tail rows contain the real data.
            if len(table) > 1:
                tail = table[1:]  # drop the repeat-header row
                cont_name, cont_cfg, force_h0 = find_continuation_config(tail, tables_config)
                if cont_name:
                    cont_maps_to = cont_cfg.get("maps_to", cont_name)
                    data = _apply_continuation(tail, cont_cfg, force_h0, cont_maps_to, result)
                    if data:
                        log(f"    INFO: Continuation data rescued from page-repeat table "
                            f"'{cont_name}' ({len(data)} rows → {cont_maps_to})", also_print=False)
            continue

        table_cfg = tables_config.get(table_name, {})
        # Try prefix matching if exact name not found (e.g. "COMMODITIES - 1 Total Lines")
        if not table_cfg:
            for cfg_name, cfg in tables_config.items():
                if cfg.get("match_prefix") and table_name.startswith(cfg["match_prefix"]):
                    table_cfg = cfg
                    break
        table_type = table_cfg.get("type", "unknown")

        if table_type == "ignore":
            continue

        maps_to = table_cfg.get("maps_to", table_name)

        if table_type == "key_value":
            data = process_key_value(table, table_cfg)
            result[maps_to] = data

        elif table_type == "data_table":
            data = process_data_table(table, table_cfg)
            if maps_to == "commodities":
                result.setdefault("commodities", []).extend(data)
            else:
                result[maps_to] = data

        elif table_type == "list":
            result[maps_to] = process_list(table, table_cfg)

        elif table_type == "certificate_fields":
            result.update(process_certificate_fields(table, table_cfg))

        elif table_type == "paired_sites":
            data = process_paired_sites(table, table_cfg)
            result.setdefault(maps_to, []).extend(data)

        elif table_type == "location_cells":
            data = process_location_cells(table, table_cfg)
            result.setdefault(maps_to, []).extend(data)

        elif table_type == "split_data_table":
            data = process_split_data_table(table, table_cfg)
            result.update(data)

        elif table_type == "049m_producer":
            data = process_049m_producer_from_page(page, table_cfg)
            result[maps_to] = data

        elif table_type == "049m_production_sites":
            data = process_049m_production_sites(table, table_cfg)
            result.setdefault(maps_to, []).extend(data)

        elif table_type == "049m_storage":
            data = process_049m_storage(table, table_cfg)
            result.setdefault(maps_to, []).extend(data)

        elif table_type == "049m_commodities":
            data = process_049m_commodities(table, table_cfg)
            result.setdefault(maps_to, []).extend(data)

        elif table_type == "049m_second_certs":
            data = process_049m_second_certs(table, table_cfg)
            for key, rows in data.items():
                result.setdefault(key, []).extend(rows)

        elif table_type == "049m_counties_and_certs":
            # Combo table: row 0 = overflow counties, rest = second certificates
            if table[0]:
                for cell in table[0]:
                    if cell:
                        name = normalize(cell).strip()
                        if name:
                            result.setdefault("authorized_counties", []).append(name)
            data = process_049m_second_certs(table, table_cfg)
            for key, rows in data.items():
                result.setdefault(key, []).extend(rows)

        elif table_type == "049m_auth_counties":
            data = process_049m_auth_counties(table, table_cfg)
            result.setdefault(maps_to, []).extend(data)

        elif table_type == "049m_auth_reps":
            data = process_049m_auth_reps(table, table_cfg)
            result.setdefault(maps_to, []).extend(data)

        elif table_type == "049m_mega_table":
            data = process_049m_mega_table(table, table_cfg)
            for key, rows in data.items():
                if isinstance(rows, list):
                    result.setdefault(key, []).extend(rows)
                else:
                    result[key] = rows

        else:
            # Not a named table — try table_patterns before flagging as unknown
            pattern_matched = False
            for _pname, pattern_cfg in county_config.get("table_patterns", {}).items():
                if not match_table_pattern(table, pattern_cfg):
                    continue
                pattern_matched = True
                ptype   = pattern_cfg.get("type", "")
                maps_to = pattern_cfg.get("maps_to", _pname)
                if ptype == "paired_sites":
                    data = process_paired_sites(table, pattern_cfg)
                    result.setdefault(maps_to, []).extend(data)
                elif ptype == "location_cells":
                    data = process_location_cells(table, pattern_cfg)
                    result.setdefault(maps_to, []).extend(data)
                elif ptype == "list":
                    result[maps_to] = process_list(table, pattern_cfg)
                elif ptype == "split_data_table":
                    result.update(process_split_data_table(table, pattern_cfg))
                elif ptype == "data_table":
                    data = process_data_table(table, pattern_cfg)
                    if maps_to == "commodities":
                        result.setdefault("commodities", []).extend(data)
                    else:
                        result.setdefault(maps_to, []).extend(data)
                elif ptype == "positional":
                    data = process_positional_table(table, pattern_cfg)
                    if data:
                        result.setdefault(maps_to, []).extend(data)
                        log(f"    INFO: Positional table '{_pname}' matched"
                            f" ({len(data)} rows → {maps_to})", also_print=False)
                elif ptype == "049m_counties_and_certs":
                    # Combo table: row 0 = overflow counties, rest = second certificates
                    if table[0]:
                        for cell in table[0]:
                            if cell:
                                name = normalize(cell).strip()
                                if name:
                                    result.setdefault("authorized_counties", []).append(name)
                    data = process_049m_second_certs(table, pattern_cfg)
                    for key, rows in data.items():
                        result.setdefault(key, []).extend(rows)
                elif ptype == "ignore":
                    pass
                break

            if not pattern_matched:
                # Try page-break continuation: table title was on prior page.
                cont_name, cont_cfg, force_h0 = find_continuation_config(table, tables_config)
                if cont_name:
                    cont_maps_to = cont_cfg.get("maps_to", cont_name)
                    data = _apply_continuation(table, cont_cfg, force_h0, cont_maps_to, result)
                    log(f"    INFO: Page-break continuation for '{cont_name}'"
                        f" ({len(data)} rows → {cont_maps_to})", also_print=False)
                else:
                    # Truly unknown — form may have changed
                    log(f"    WARNING: Unknown table '{table_name}' — not in county config. "
                        f"Form may have changed. Review county_configs/{county_config.get('county','?')}.json")
                    header_idx = find_header_row(table)
                    if header_idx >= len(table):
                        continue
                    headers = [normalize(h) for h in table[header_idx]]
                    rows = []
                    for row in table[header_idx + 1:]:
                        if not any(c and normalize(c) for c in row):
                            continue
                        rows.append({h: normalize(v) for h, v in zip(headers, row) if h})
                    if rows:
                        result[table_name] = rows

    return result

def extract_page_text_based(page, county_config):
    """Text-based extraction for counties like San Diego where data is in free text."""
    result = {}
    page_text = page.extract_text() or ""
    tables = page.extract_tables()

    # Certificate header fields
    cert_fields = extract_cert_fields_from_text(page_text, county_config)
    if cert_fields:
        result.update(cert_fields)

    # Producer fields
    producer = extract_producer_fields_from_text(page_text, county_config)
    if producer:
        result["producer"] = producer

    # Text section parsing (authorized counties, reps, producers_i_sell_for, etc.)
    sections = extract_text_sections(page_text, county_config)
    for key, value in sections.items():
        result[key] = value

    # Character-position commodity extraction (Santa Barbara style)
    if county_config.get("commodity_column_boundaries"):
        commodities = extract_commodities_by_char_position(page, county_config)
        if commodities:
            result.setdefault("commodities", []).extend(commodities)

    # Table pattern matching
    table_patterns = county_config.get("table_patterns", {})
    for table in tables:
        if not table:
            continue
        # Check each pattern
        for _pattern_name, pattern_cfg in table_patterns.items():
            if not match_table_pattern(table, pattern_cfg):
                continue
            ptype   = pattern_cfg.get("type", "")
            maps_to = pattern_cfg.get("maps_to", _pattern_name)
            if ptype == "production_sites_cells":
                data = process_production_sites_cells(table, pattern_cfg)
                if data:
                    result.setdefault("production_sites", []).extend(data)
            elif ptype == "positional":
                data = process_positional_table(table, pattern_cfg)
                if data:
                    if maps_to == "commodities":
                        result.setdefault("commodities", []).extend(data)
                    else:
                        result.setdefault(maps_to, []).extend(data)
            break  # matched — no need to check more patterns for this table

    # Ensure all standard list fields are present (empty arrays if not extracted)
    for list_field in ("production_sites", "storage_locations", "authorized_counties",
                       "producers_i_sell_for", "producers_selling_for_me",
                       "authorized_representatives", "commodities"):
        result.setdefault(list_field, [])

    return result

def extract_page(page, county_config):
    if county_config.get("extraction_method") == "text_based":
        return extract_page_text_based(page, county_config)
    return extract_page_table_based(page, county_config)

# ── CONFIG GENERATION ────────────────────────────────────────────

def generate_county_config(county, pdf_path, page1_text, all_table_names):
    existing = {}
    for f in sorted(COUNTY_CONFIGS_FOLDER.glob("*.json")):
        existing[f.stem] = json.loads(f.read_text())

    prompt = f"""You are creating a county config for California Certified Producer's Certificates
issued by {county} County.

All 58 California counties collect the same information by state law. Only the labels differ.

TARGET STANDARD SCHEMA:
{json.dumps(STANDARD_SCHEMA, indent=2)}

EXISTING COUNTY CONFIGS (learn from these patterns):
{json.dumps(existing, indent=2) if existing else "None yet."}

TABLE NAMES FOUND IN THIS DOCUMENT (these are the first cell of each table's first row):
{json.dumps(all_table_names, indent=2)}

PAGE 1 TEXT (first 2000 characters):
{page1_text[:2000]}

Generate a config JSON with this exact structure:
{{
  "county": "{county}",
  "generated_date": "{datetime.now().strftime('%Y-%m-%d')}",
  "page_repeat_prefixes": ["<text that page-repeat summary tables start with>"],
  "certificate_fields_from_text": {{
    "<exact label text from document>": "<standard field name>",
    ...
  }},
  "tables": {{
    "<exact table name>": {{
      "type": "<ignore|key_value|data_table|list|certificate_fields>",
      "maps_to": "<standard collection name e.g. commodities, production_sites, producer>",
      "field_map or column_map": {{
        "<exact label/column from document>": "<standard field name>"
      }}
    }},
    ...
  }},
  "notes": "..."
}}

Table types:
- ignore: skip this table (boilerplate, signatures, page headers)
- key_value: two columns, left=label right=value (e.g. Producer Information)
- data_table: title row + header row + data rows (e.g. Commodities, Production Sites)
- list: single column or comma-separated values (e.g. authorized counties)
- certificate_fields: borderless label:value table for certificate header info

Output only raw JSON, no commentary.
"""
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = clean_json_response(response.content[0].text)
    return json.loads(raw)

# ── XLS/XLSX EXTRACTION ──────────────────────────────────────────

def excel_serial_to_date(serial):
    """Convert Excel serial number to date string MM/DD/YYYY."""
    if not serial:
        return ""
    try:
        serial = float(serial)
        dt = xlrd.xldate_as_datetime(int(serial), 0)
        return dt.strftime("%m/%d/%Y")
    except (ValueError, TypeError, OverflowError):
        return str(serial)

def detect_xls_format(file_path):
    """Auto-detect which XLS config to use based on sheet name and structure.
    Returns (config_dict, config_path) or (None, None)."""
    suffix = file_path.suffix.lower()

    if suffix == ".xls":
        if not xlrd:
            log("  ERROR: xlrd not installed. Run: pip install xlrd")
            return None, None
        wb = xlrd.open_workbook(str(file_path), on_demand=True)
        sheet_names = wb.sheet_names()
        # Check for Access export format (sheet named "Crops")
        if "Crops" in sheet_names:
            sheet = wb.sheet_by_name("Crops")
            row1 = [str(sheet.cell(1, c).value) if c < sheet.ncols else ""
                    for c in range(min(sheet.ncols, 6))]
            if "CERTNUMBER" in row1 or "Combo239" in row1:
                config_path = COUNTY_CONFIGS_FOLDER / "VENTURA_ACCESS_EXPORT.json"
                if config_path.exists():
                    return json.loads(config_path.read_text()), config_path
        # Fallback: try county-based config
        return None, None

    elif suffix == ".xlsx":
        if not openpyxl:
            log("  ERROR: openpyxl not installed. Run: pip install openpyxl")
            return None, None
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        # Check for state form format (sheet named "PrintCPCertificate")
        if "PrintCPCertificate" in sheet_names:
            sheet = wb["PrintCPCertificate"]
            # Read cell that should contain certificate info block
            cell_val = sheet.cell(5, 21).value  # row 4, col 20 (0-indexed) → openpyxl row 5, col 21
            wb.close()
            if cell_val and "ISSUING COUNTY" in str(cell_val):
                # Detect county from certificate block
                m = re.search(r'ISSUING COUNTY[:\s]+([A-Za-z][A-Za-z\s]+?)(?:\n|$)', str(cell_val), re.IGNORECASE)
                county = m.group(1).strip().upper() if m else "UNKNOWN"
                # Try county-specific state form config
                config_path = COUNTY_CONFIGS_FOLDER / f"{county}_51-049_Rev_10-24.json"
                if config_path.exists():
                    return json.loads(config_path.read_text()), config_path
                # Try LA config as fallback (same state form)
                config_path = COUNTY_CONFIGS_FOLDER / "LOS ANGELES_51-049_Rev_10-24.json"
                if config_path.exists():
                    config = json.loads(config_path.read_text())
                    config["format_type"] = "xlsx_state_form"
                    return config, config_path
        wb.close()
        return None, None

    return None, None

def extract_xls_access_export(file_path, config):
    """Extract CPC data from Ventura-style Access export XLS file.
    All certificate data is in row 0, commodities in rows 4+."""
    result = {}
    skip_val = config.get("skip_value", "N/A")
    col_map = config.get("header_column_map", {})
    date_cols = config.get("date_columns", [])

    wb = xlrd.open_workbook(str(file_path), on_demand=True)
    sheet_name = config.get("sheet_name", "Crops")
    sheet = wb.sheet_by_name(sheet_name)

    # Extract header fields from row 0
    producer = {}
    for col_str, field_name in col_map.items():
        col = int(col_str)
        if col >= sheet.ncols:
            continue
        value = sheet.cell(0, col).value
        if value is None or str(value).strip().upper() == skip_val:
            value = ""
        else:
            value = str(value).strip()
            # Convert Excel serial dates
            if col in date_cols and value:
                value = excel_serial_to_date(value)
            # Clean up numeric strings (e.g. "30.0" → "30")
            if value.endswith(".0") and value[:-2].isdigit():
                value = value[:-2]

        if field_name.startswith("producer."):
            producer[field_name.replace("producer.", "")] = value
        else:
            result[field_name] = value

    # Clean zip code trailing spaces/dashes
    if "zip_code" in producer:
        producer["zip_code"] = producer["zip_code"].rstrip("- ")
    if producer:
        result["producer"] = producer

    # Extract production sites from cols 14-18
    site_cols = config.get("production_site_columns", [14, 15, 16, 17, 18])
    sites = []
    for i, col in enumerate(site_cols, start=1):
        if col >= sheet.ncols:
            continue
        value = str(sheet.cell(0, col).value or "").strip()
        if value and value.upper() != skip_val:
            sites.append({
                "site_number": str(i).zfill(2),
                "description": value
            })
    if sites:
        result["production_sites"] = sites

    # Extract storage locations from cols 20, 22
    storage_cols = config.get("storage_location_columns", [20, 22])
    locations = []
    for i, col in enumerate(storage_cols):
        if col >= sheet.ncols:
            continue
        value = str(sheet.cell(0, col).value or "").strip()
        if value and value.upper() != skip_val:
            loc_id = chr(65 + i)  # A, B, C...
            locations.append({
                "location_id": loc_id,
                "address": value
            })
    if locations:
        result["storage_locations"] = locations

    # Extract commodities from rows 4+
    comm_col_map = config.get("commodity_column_map", {})
    header_rows = set(config.get("commodity_header_rows", [3, 25]))
    start_row = config.get("commodity_start_row", 4)
    commodities = []

    for row_idx in range(start_row, sheet.nrows):
        # Skip commodity header repeat rows
        if row_idx in header_rows:
            continue
        # Check if row has commodity data (col 25 = crop name)
        crop_col = 25
        for col_str in comm_col_map:
            if comm_col_map[col_str] == "commodity":
                crop_col = int(col_str)
                break
        crop_value = str(sheet.cell(row_idx, crop_col).value or "").strip()
        if not crop_value:
            continue

        entry = {}
        for col_str, field_name in comm_col_map.items():
            col = int(col_str)
            if col >= sheet.ncols:
                continue
            if field_name == "row_number":
                continue  # skip the row counter column
            value = str(sheet.cell(row_idx, col).value or "").strip()
            # Clean non-breaking spaces
            value = value.replace("\xa0", " ").strip()
            entry[field_name] = value
        # Ensure all standard commodity fields are present (empty if not in this format)
        for std_field in ("site", "commodity", "variety", "amount_grown",
                          "est_production", "harvest_season",
                          "season_altering_device", "months_in_storage"):
            entry.setdefault(std_field, "")
        commodities.append(entry)

    if commodities:
        result["commodities"] = commodities

    # Ensure all standard schema fields are present (empty if not in this format)
    result.setdefault("authorized_counties", [])
    result.setdefault("producers_i_sell_for", [])
    result.setdefault("producers_selling_for_me", [])
    result.setdefault("authorized_representatives", [])
    result.setdefault("commodities", [])
    result.setdefault("amended_date", "")
    result.setdefault("certified_copies_made", "")

    log(f"  Extracted {len(commodities)} commodities from Access export")
    return result

def extract_xlsx_state_form(file_path, config):
    """Extract CPC data from state-form XLSX file (e.g. Shigeru Nursery).
    Form-based layout with labeled rows and sections."""
    result = {}

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheet_name = config.get("sheet_name", "PrintCPCertificate")
    sheet = wb[sheet_name]

    # Helper: get cell value (openpyxl is 1-indexed)
    def cell(row, col):
        v = sheet.cell(row + 1, col + 1).value
        if v is None:
            return ""
        return str(v).strip()

    # Extract certificate fields from multi-line text block
    cert_cfg = config.get("certificate_cell", {})
    cert_row = cert_cfg.get("row", 4)
    cert_col = cert_cfg.get("col", 19)
    cert_text = cell(cert_row, cert_col)
    if cert_text:
        for pattern_label, field_name in config.get("certificate_field_patterns", {}).items():
            m = re.search(re.escape(pattern_label) + r'[:\s]+([^\n]*)', cert_text, re.IGNORECASE)
            if m:
                value = m.group(1).strip()
                result[field_name] = value

    # Process each section
    sections = config.get("sections", {})

    # Build a map of row contents for quick section detection
    max_row = sheet.max_row
    max_col = sheet.max_column
    row_cache = {}
    for r in range(max_row):
        for c in range(max_col):
            v = cell(r, c)
            if v:
                row_cache.setdefault(r, {})[c] = v

    # Find section start rows by searching for marker text
    # Use exact match (cell value starts with marker) to avoid false positives
    # e.g. "Commodities" must not match "Authorized Counties Where Commodities May Be Sold"
    section_starts = {}
    for section_name, section_cfg in sections.items():
        marker = section_cfg.get("marker", "")
        if not marker:
            continue
        for r, cols in row_cache.items():
            for c, v in cols.items():
                if v.strip() == marker or v.strip().startswith(marker + "\n"):
                    section_starts[section_name] = r
                    break
            if section_name in section_starts:
                break

    # Sort sections by row position
    ordered_sections = sorted(section_starts.items(), key=lambda x: x[1])

    for idx, (section_name, start_row) in enumerate(ordered_sections):
        section_cfg = sections[section_name]
        section_type = section_cfg.get("type", "")
        # Determine end row (next section or end of sheet)
        end_row = ordered_sections[idx + 1][1] if idx + 1 < len(ordered_sections) else max_row

        if section_type == "key_value_rows":
            label_col = section_cfg.get("label_col", 1)
            value_col = section_cfg.get("value_col", 14)
            field_map = section_cfg.get("field_map", {})
            data = {}
            for r in range(start_row + 1, end_row):
                label = cell(r, label_col)
                value = cell(r, value_col)
                if label in field_map:
                    # Handle multi-line address
                    clean_val = re.sub(r'\s+', ' ', value).strip() if value else ""
                    data[field_map[label]] = clean_val
            if data:
                result[section_name] = data

        elif section_type == "data_table_rows":
            header_labels = section_cfg.get("header_labels", {})
            col_positions = {int(k): v for k, v in header_labels.items()}
            # data_offset: 2 = marker row + header row + data (default)
            #              1 = marker IS the header, data starts next row
            data_offset = section_cfg.get("data_offset", 2)
            data_start = start_row + data_offset
            rows = []
            for r in range(data_start, end_row):
                entry = {}
                has_data = False
                for col_pos, field_name in col_positions.items():
                    value = cell(r, col_pos)
                    value = value.replace("\xa0", " ").strip() if value else ""
                    entry[field_name] = value
                    if value:
                        has_data = True
                if has_data:
                    rows.append(entry)
            if rows:
                result[section_name] = rows

        elif section_type == "comma_list":
            data_col = section_cfg.get("data_col", 3)
            items = []
            for r in range(start_row + 1, end_row):
                value = cell(r, data_col)
                if value:
                    for item in re.split(r'[,\n]+', value):
                        item = item.strip()
                        if item:
                            items.append(item)
            if items:
                result[section_name] = items

    wb.close()
    commodity_count = len(result.get("commodities", []))
    log(f"  Extracted {commodity_count} commodities from state form XLSX")
    return result

# ── MAIN ─────────────────────────────────────────────────────────

pdf_folder = Path("./certificates")
page_output_folder = Path("./page_output")
page_output_folder.mkdir(exist_ok=True)

all_files = sorted(
    list(pdf_folder.glob("*.pdf")) +
    list(pdf_folder.glob("*.xls")) +
    list(pdf_folder.glob("*.xlsx")),
    key=lambda f: f.name.lower()
)
# Exclude subdirectories' files — only top-level certificates folder
all_files = [f for f in all_files if f.parent == pdf_folder]

if not all_files:
    print("No PDF/XLS/XLSX files found in ./certificates")
    exit()

print("Which file to convert?")
print("  0 - Process ALL files")
for i, f in enumerate(all_files, start=1):
    print(f"  {i} - {f.name}")

choice = input("Enter number: ").strip()
if choice == "0":
    files = all_files
elif choice.isdigit() and 1 <= int(choice) <= len(all_files):
    files = [all_files[int(choice) - 1]]
else:
    print("Invalid choice. Exiting.")
    exit()

run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
log(f"\n{'='*60}")
log(f"CONVERSION RUN: {run_time}")
log(f"{'='*60}")

for input_file in files:
    log(f"\nFILE: {input_file.name}")
    file_ext = input_file.suffix.lower()

    # ── XLS/XLSX PATH ──────────────────────────────────────────
    if file_ext in (".xls", ".xlsx"):
        xls_config, config_file = detect_xls_format(input_file)
        if not xls_config:
            log(f"  ERROR: No config found for this spreadsheet format. Skipping.")
            continue
        log(f"  Format: {xls_config.get('format_type', 'unknown')}")
        log(f"  Config: {config_file.name}")

        format_type = xls_config.get("format_type", "")
        if format_type == "xls_access_export":
            data = extract_xls_access_export(input_file, xls_config)
        elif format_type == "xlsx_state_form":
            data = extract_xlsx_state_form(input_file, xls_config)
        else:
            log(f"  ERROR: Unknown format_type '{format_type}'. Skipping.")
            continue

        # XLS files produce a single output (no pages)
        file_page_folder = page_output_folder / input_file.stem
        file_page_folder.mkdir(exist_ok=True)
        page_file = file_page_folder / "page_001.json"
        page_file.write_text(json.dumps(data, indent=2, ensure_ascii=False))
        total_commodities = len(data.get("commodities", []))
        log(f"  TOTAL: {total_commodities} commodities")
        log(f"\n  Done. Run cpcMerge.py to combine into final JSON.")
        continue

    # ── PDF PATH (existing logic) ──────────────────────────────
    pdf_file = input_file

    with pdfplumber.open(pdf_file) as pdf:
        page1_text = pdf.pages[0].extract_text() or ""
        page1_tables = pdf.pages[0].extract_tables()
        total_pages = len(pdf.pages)

        # Collect all unique table names across the whole document
        all_table_names = set()
        for page in pdf.pages:
            for table in page.extract_tables():
                name = get_table_name(table)
                if name:
                    all_table_names.add(name)

    log(f"  Total pages: {total_pages}")

    county = detect_county(page1_text, page1_tables)
    if not county:
        log(f"  WARNING: Could not auto-detect county.")
        county = "UNKNOWN"
    log(f"  Issuing county: {county}")

    form_revision = detect_form_revision(page1_text)
    if form_revision:
        log(f"  Form revision: {form_revision}")
    else:
        log(f"  Form revision: not detected")

    county_config, config_file = find_county_config(county, form_revision)

    if county_config:
        log(f"  County config: loaded ({config_file.name})")
        if form_revision and not county_config.get("form_revision"):
            county_config["form_revision"] = form_revision
            config_file.write_text(json.dumps(county_config, indent=2, ensure_ascii=False))
            log(f"  Form revision recorded in config: {form_revision}")
    else:
        if form_revision:
            rev_part = revision_to_filename_part(form_revision)
            new_config_file = COUNTY_CONFIGS_FOLDER / f"{county}_{rev_part}.json"
        else:
            new_config_file = COUNTY_CONFIGS_FOLDER / f"{county}.json"

        log(f"  County config: NOT FOUND — generating for {county} County (one-time AI call)...")
        county_config = generate_county_config(county, pdf_file, page1_text, sorted(all_table_names))
        if form_revision:
            county_config["form_revision"] = form_revision
        new_config_file.write_text(json.dumps(county_config, indent=2, ensure_ascii=False))
        config_file = new_config_file
        log(f"  County config: CREATED — saved to county_configs/{config_file.name}")
        log(f"  NOTE: Review county_configs/{config_file.name} before processing more {county} CPCs.")

    pdf_page_folder = page_output_folder / pdf_file.stem
    pdf_page_folder.mkdir(exist_ok=True)

    page_commodity_counts = []

    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            log(f"  Page {page_num}/{total_pages}: extracting...")
            data = extract_page(page, county_config)
            page_file = pdf_page_folder / f"page_{page_num:03d}.json"
            page_file.write_text(json.dumps(data, indent=2, ensure_ascii=False))
            count = len(data.get("commodities", []))
            page_commodity_counts.append(count)

    log(f"\n  --- COMMODITY COUNT PER PAGE ---")
    total_commodities = 0
    for i, count in enumerate(page_commodity_counts, start=1):
        log(f"    Page {i:02d}: {count} commodities")
        total_commodities += count
    log(f"    TOTAL: {total_commodities} commodities")
    log(f"\n  Done. Run cpcMerge.py to combine into final JSON.")

log(f"\n{'='*60}")
log(f"Run complete: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
log(f"{'='*60}\n")
